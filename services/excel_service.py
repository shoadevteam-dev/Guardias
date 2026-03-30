"""
Servicio de exportación a Excel
"""
import io
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from models.models import Guardia, Persona
from services.guardias_service import calcular_retenes_por_mes
from services.consultas import obtener_rango_mes, obtener_guardias_mes, formatear_nombre


def exportar_guardias_excel(mes, anio):
    """Exporta las guardias del mes a un archivo Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Guardias {mes}-{anio}"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    inicio_mes, fin_mes = obtener_rango_mes(mes, anio)

    # Obtener guardias
    guardias = obtener_guardias_mes(mes, anio)
    guardias_dict = {g.fecha: g for g in guardias}

    # Calcular retenes usando la misma lógica que la API
    reten_por_fecha, reten_contador = calcular_retenes_por_mes(mes, anio)

    # Nombres
    nombres_meses = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
        7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    # weekday() de Python: 0=Lunes, 1=Martes, ..., 6=Domingo
    nombres_dias = {
        0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves',
        4: 'Viernes', 5: 'Sábado', 6: 'Domingo'
    }

    # Encabezados
    headers = ['Fecha', 'Día', 'Mes', 'Informático de Guardia', 'Retén']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # Datos
    row = 2
    fecha_actual = inicio_mes

    while fecha_actual <= fin_mes:
        guardia = guardias_dict.get(fecha_actual.date())

        if guardia:
            persona = Persona.query.get(guardia.persona_id)
            nombre_persona = 'PAC ' + formatear_nombre(persona.nombre) if persona else 'N/A'

            reten_persona_id = reten_por_fecha.get(fecha_actual.date())
            if reten_persona_id:
                persona_reten = Persona.query.get(reten_persona_id)
                reten = 'PAC ' + formatear_nombre(persona_reten.nombre) if persona_reten else 'SIN RETÉN'
            else:
                reten = 'SIN RETÉN'
        else:
            nombre_persona = 'SIN ASIGNAR'
            reten = 'SIN RETÉN'

        ws.cell(row=row, column=1, value=fecha_actual.day).border = border
        ws.cell(row=row, column=2, value=nombres_dias[fecha_actual.weekday()]).border = border
        ws.cell(row=row, column=3, value=nombres_meses.get(mes, f'Mes {mes}')).border = border
        ws.cell(row=row, column=4, value=nombre_persona).border = border
        ws.cell(row=row, column=5, value=reten).border = border

        for col in range(1, 6):
            ws.cell(row=row, column=col).alignment = center_alignment

        row += 1
        fecha_actual += timedelta(days=1)

    # Obtener personas activas para el resumen
    personas_activas = Persona.query.filter_by(activo=True).all()

    # Hoja de resumen
    _agregar_resumen_excel(
        wb, personas_activas, guardias, reten_contador,
        header_font, header_fill, border, center_alignment
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _agregar_resumen_excel(wb, personas, guardias, reten_contador,
                           header_font, header_fill, border, center_alignment):
    """Agrega hoja de resumen al Excel - Solo guardias del mes"""
    ws2 = wb.create_sheet(title="Resumen Guardias Mes")

    resumen_headers = ['Persona', 'Guardias este Mes', 'Retenes este Mes']
    for col, header in enumerate(resumen_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # Calcular estadísticas - solo guardias del mes
    resumen_dict = {
        p.id: {'nombre': 'PAC ' + formatear_nombre(p.nombre), 'guardias': 0, 'retenes': reten_contador.get(p.id, 0)}
        for p in personas
    }

    for g in guardias:
        if g.persona_id in resumen_dict:
            resumen_dict[g.persona_id]['guardias'] += 1

    # Escribir datos - ordenado por cantidad de guardias (mayor a menor)
    row = 2
    for datos in sorted(
        resumen_dict.values(),
        key=lambda x: x['guardias'],
        reverse=True
    ):
        ws2.cell(row=row, column=1, value=datos['nombre']).border = border
        ws2.cell(row=row, column=2, value=datos['guardias']).border = border
        ws2.cell(row=row, column=3, value=datos['retenes']).border = border

        for col in range(1, 4):
            ws2.cell(row=row, column=col).alignment = center_alignment
        row += 1

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 18
